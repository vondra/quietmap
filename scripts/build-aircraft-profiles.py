#!/usr/bin/env python3
"""Generate engine/noise-compute/src/emission/profiles_generated.rs.

Two inputs:

  1. EASA ANP v2.3 CSVs (`--anp <DIR>`):
     - ANP2.3_Aircraft.csv  : ACFT_ID → (NPD_ID, Engine Type, Number Of Engines,
                              Weight Class, Lateral Directivity Identifier)
     - ANP2.3_NPD_data.csv  : (NPD_ID, Op Mode) → 2..13 SEL rows, each with
                              Power Setting + 10 distance SEL values

  2. Pinned global traffic counts (`--counts <PATH>`):
     - JSON dict {ICAO_typecode → segment_count} from running aircraft
       extraction across the global ADS-B archive. Snapshot pinned in repo
       at scripts/aircraft-profiles-counts.json. SHA stamped into the
       generated banner so any drift between snapshot and generated output
       is visible to reviewers.

Output (stdout): Rust source for `profiles_generated.rs`.

Class scheme: **Voronoi assignment over NUM_CLASSES anchor classes (14 traffic-ranked + pinned).**

Each anchor is the exact NPD vector of one dominant profile (top-K
traffic per Installation). Every other profile is assigned to its
nearest anchor by L∞ distance, constrained to the same Installation.
Helicopters pinned to their own class (different physics, single
ANP-equivalent reference NPD shared by all rotorcraft).

Per-Installation anchor counts: 9 Wing + 2 Fuselage + 2 Propeller + 1
Helicopter = 14 classes. Computed from
scripts/aircraft-profiles-counts.json (4.885 G global segments,
51 363 R4 cells); weighted-avg per-segment acoustic error = 0.76 dB
across all 124 profiles, well under Doc 29 measurement uncertainty.
Wing=9 (vs the original Wing=7 design at c93f5f4) restores B789 +
A319 anchors after the c76cefa A20N/A21N v9 split displaced them.

FALLBACK NPD is recomputed at generation time as the traffic-weighted
energy-mean of every non-FALLBACK profile, so unmapped typecodes get
the best estimate of the global average aircraft instead of a B738
proxy (which over-counted by ~2 dB at dep@1k).

Power Setting selection: lowest for Approach (idle/low-thrust), highest
for Departure (max takeoff thrust).

Hand-curated `ICAO_TYPECODE_TO_ACFT_ID` maps ICAO 4-letter typecodes to
ANP ACFT_IDs. Non-ANP types (GA piston, helicopters, ANP-missing neos)
fall back to manual placeholder profiles derived from Doc 29 Vol 3
reference curves (PROP/Helo) plus engineering judgment.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import sys
from dataclasses import dataclass, field
from pathlib import Path

# ─── Hand-curated ICAO typecode → ANP ACFT_ID ────────────────────────────
# Order = profile_idx in the emitted PROFILES array. FALLBACK must be last
# (its index = NUM_PROFILES - 1). Each entry is
# (typecode, ACFT_ID_or_None, manual_class_or_None). Manual entries skip
# ANP NPD lookup and use a placeholder from MANUAL_PROFILES.
ICAO_TYPECODE_TO_ACFT_ID: list[tuple[str, str | None, str | None]] = [
    # ── narrowbody jets (high-bypass) ─────────────────────────────────────
    ("B738", "737800", None),            # CFM56-7B26
    ("B739", "737800", None),            # 737-900 → use -800 nearest
    ("B737", "7373B2", None),            # 737-300 / CFM56-3B-2
    ("B734", "737400", None),
    ("B735", "737500", None),
    ("B733", "7373B2", None),            # 737-300 nearest
    ("B736", "737400", None),
    ("B38M", "7378MAX", None),           # 737 MAX 8 / LEAP-1B-27
    ("B39M", "7378MAX", None),           # 737 MAX 9 nearest
    ("B37M", "7378MAX", None),           # 737 MAX 7 nearest
    ("A320", "A320-232", None),          # V2527-A5
    # A20N: prefer EASA ANP v9 dedicated A320-270N (PW1100G GTF) — falls back
    # to v2.3 A320-232 (V2527-A5) if v9 supplement absent.
    ("A20N", "A320-270N", None),         # neo PW1100G (v9); fallback A320-232
    ("A319", "A319-131", None),          # V2522-A5
    ("A19N", "A319-131", None),          # A319neo nearest classic
    ("BCS3", "A319-131", None),          # A220-300 nearest narrowbody
    ("BCS1", "A319-131", None),          # A220-100 nearest narrowbody
    # ── stretched narrowbody / 757 ────────────────────────────────────────
    ("A321", "A321-232", None),          # V2530
    ("A21N", "A321-270N", None),         # neo PW1100G (v9); fallback A321-232
    ("B752", "757PW", None),             # PW2037
    ("B753", "757PW", None),             # 757-300 nearest -200
    # ── widebody 2-engine ─────────────────────────────────────────────────
    ("B772", "777200", None),            # GE90-76B
    ("B773", "777300", None),            # Trent 892
    ("B77W", "7773ER", None),            # 777-300ER / GE90-115B
    ("B77L", "777200", None),            # 777-200LR nearest
    ("B77F", "777300", None),            # 777F nearest pax
    ("B788", "7878R", None),             # 787-8 / Trent 1000
    ("B789", "7878R", None),             # 787-9 nearest -8
    ("B78X", "7878R", None),             # 787-10 nearest -8
    ("A332", "A330-301", None),          # CF6-80
    ("A333", "A330-343", None),          # Trent 772B
    ("A338", "A330-343", None),          # A330-800neo nearest classic
    ("A339", "A330-941", None),          # A330-900neo dedicated (v9); fallback A330-343
    ("A359", "A350-941", None),          # Trent XWB-84
    ("A35K", "A350-1041", None),         # A350-1000 dedicated (v9); fallback A350-941
    ("A306", "A330-301", None),          # A300 nearest A330
    ("A310", "A330-301", None),          # A310 nearest A330
    ("B763", "767300", None),            # CF6-80A
    ("B764", "767400", None),            # CF6-80C2
    # ── widebody 3-engine ─────────────────────────────────────────────────
    ("MD11", "MD11GE", None),            # CF6-80C2
    ("DC10", "DC1030", None),
    ("L101", "L1011", None),             # Lockheed L-1011 / RB211-22B
    # ── widebody 4-engine ─────────────────────────────────────────────────
    ("B744", "747400", None),            # PW4056
    ("B748", "7478", None),              # 747-8F / GEnx-2B67
    ("B741", "747100", None),
    ("B742", "747200", None),
    ("A342", "A340-211", None),
    ("A343", "A340-211", None),
    ("A346", "A340-642", None),          # Trent 556
    ("A388", "A380-841", None),          # Trent 970
    ("IL76", "747200", None),            # IL-76 nearest 747-200 (heavy 4-eng wing)
    # ── regional jets / fuselage-mounted ─────────────────────────────────
    ("E170", "EMB170", None),            # ERJ170 / CF34-8E
    ("E75L", "EMB175", None),            # ERJ175
    ("E75S", "EMB175", None),
    ("E190", "EMB190", None),            # ERJ190 / CF34-10E
    ("E195", "EMB195", None),            # ERJ195
    ("E290", "ERJ190-300", None),        # E190-E2 dedicated (v9); fallback EMB190
    ("E295", "ERJ190-400", None),        # E195-E2 dedicated (v9); fallback EMB195
    ("CRJ2", "CL600", None),
    ("CRJ7", "CL601", None),
    ("CRJ9", "CL601", None),
    ("EMJ", "EMB145", None),             # E145
    # ── business jets ─────────────────────────────────────────────────────
    ("CL60", "CL601", None),             # Challenger 600 / 601 — pick CF34 over older ALF502L
    ("C56X", "CIT3", None),              # Citation Excel
    ("C680", "CIT3", None),              # Citation Sovereign nearest
    ("GLEX", "CL601", None),             # Global Express nearest CF34 (BR710 isn't in ANP)
    ("GLF6", "CL601", None),             # G650 — ANP only has 1970s GII Spey (~125 dB);
                                         # CL601 CF34 is the closest MODERN turbofan (~100 dB)
    ("GLF5", "CL601", None),             # G550 — same rationale as G650
    ("FA7X", "CL601", None),             # Falcon 7X — modern triple-engine, FAL20 is 1970s Spey
    ("PC24", None, "JET_BIZ_FUSE"),      # PC-24 manual
    ("LJ60", "LEAR35", None),            # Learjet 60 nearest
    # ── turboprops large ──────────────────────────────────────────────────
    ("AT72", "DHC830", None),            # ATR-72 → Dash 8 Q400 nearest large TP
    ("AT76", "DHC830", None),
    ("AT43", "DHC8", None),              # ATR-42 → Dash 8 nearest
    ("AT45", "DHC8", None),
    ("DH8D", "DHC830", None),            # Dash 8 Q400
    ("DH8C", "DHC8", None),
    ("DH8A", "DHC8", None),
    ("DH8B", "DHC8", None),
    ("L410", "DHC6", None),              # Let L-410 → Twin Otter nearest small TP
    ("EN48", "DHC6", None),              # Embraer 110 → Twin Otter
    ("SF34", "DHC8", None),
    ("F50", "DHC830", None),             # Fokker 50
    ("F70", "EMB145", None),             # Fokker 70 (regional jet, but tiny — RJ class)
    ("JS41", "DHC8", None),              # Jetstream 41
    # ── light GA piston (manual placeholders, no ANP) ───────────────────
    ("C172", None, "PISTON_SE"),
    ("C152", None, "PISTON_SE"),
    ("C182", None, "PISTON_SE"),
    ("PA28", None, "PISTON_SE"),
    ("PA34", None, "PISTON_TWIN"),
    ("SR20", None, "PISTON_SE"),
    ("SR22", None, "PISTON_SE"),
    ("DA40", None, "PISTON_SE"),
    ("DA42", None, "PISTON_TWIN"),
    ("P28A", None, "PISTON_SE"),
    ("C210", None, "PISTON_SE"),
    ("BE36", None, "PISTON_SE"),
    ("M20P", None, "PISTON_SE"),
    ("C206", None, "PISTON_SE"),
    ("PA32", None, "PISTON_SE"),
    ("PA44", None, "PISTON_TWIN"),
    ("RV7", None, "PISTON_SE"),
    ("RV8", None, "PISTON_SE"),
    # ── helicopters (manual placeholders, no ANP) ───────────────────────
    ("EC35", None, "HELICOPTER"),
    ("EC45", None, "HELICOPTER"),
    ("EC55", None, "HELICOPTER"),
    ("EC30", None, "HELICOPTER"),
    ("EC20", None, "HELICOPTER"),
    ("AS50", None, "HELICOPTER"),
    ("AS55", None, "HELICOPTER"),
    ("AS65", None, "HELICOPTER"),
    ("H500", None, "HELICOPTER"),
    ("MD52", None, "HELICOPTER"),
    ("B06", None, "HELICOPTER"),
    ("B407", None, "HELICOPTER"),
    ("B412", None, "HELICOPTER"),
    ("R22", None, "HELICOPTER"),
    ("R44", None, "HELICOPTER"),
    ("R66", None, "HELICOPTER"),
    ("S76", None, "HELICOPTER"),
    ("A109", None, "HELICOPTER"),
    ("BK17", None, "HELICOPTER"),
    ("B505", None, "HELICOPTER"),
    ("GYRO", None, "HELICOPTER"),
    # ── fallback (must be last, idx = NUM_PROFILES - 1) ──────────────────
    # ACFT_ID is a placeholder; FALLBACK NPD is overwritten with the
    # traffic-weighted energy-mean of all non-FALLBACK profiles before
    # Voronoi clustering. See `recompute_fallback_npd`.
    ("FALLBACK", "737800", None),
]


# ─── Manual placeholder profiles (for non-ANP types) ─────────────────────
# Derived from Doc 29 Vol 3 reference curves (PROP) plus engineering judgment.
# Tuple = (approach_sel[10], departure_sel[10], v_ref_kt, d_bar_m,
#          installation, is_jet). Helicopters carry Installation::Propeller
# because the Rust enum lacks a rotorcraft variant — rotor Λ behaviour is
# a TODO per SPEC.md and Propeller is the closest single-Installation fit.
MANUAL_PROFILES: dict[str, tuple[list[float], list[float], float, float, str, bool]] = {
    # Single-engine piston (C172, PA28). Quiet, ~85 dB at 200 ft, fast roll-off.
    "PISTON_SE": (
        [85.0, 80.0, 76.0, 72.0, 65.0, 58.0, 53.0, 47.0, 41.0, 35.0],
        [88.0, 83.0, 79.0, 75.0, 68.0, 61.0, 56.0, 50.0, 44.0, 38.0],
        90.0, 208.0, "Prop", False,
    ),
    # Twin-engine piston (PA34, DA42). +3 dB vs single.
    "PISTON_TWIN": (
        [88.0, 83.0, 79.0, 75.0, 68.0, 61.0, 56.0, 50.0, 44.0, 38.0],
        [91.0, 86.0, 82.0, 78.0, 71.0, 64.0, 59.0, 53.0, 47.0, 41.0],
        110.0, 220.0, "Prop", False,
    ),
    # Helicopter NPD anchored at energy-mean of 64.2M global heli events
    # (88 dB SEL @ 1000 ft) computed from scripts/aircraft-profiles-counts.json.
    # Curve shape: MV-22 helicopter-mode AAM NPD (ACRP02-44 Table 4-7,
    # https://onlinepubs.trb.org/onlinepubs/acrp/docs/ACRP02-44_FR.pdf) —
    # closest publicly-tabulated rotorcraft per-distance NPD; represents
    # heavy-rotorcraft slope. Raw MV-22 = [100.6, 97.2, 94.8, 92.3, 87.9,
    # 82.5, 78.7, 74.0, 68.0, 61.2] dB @ [200, 400, 630, 1000, 2000, 4000,
    # 6300, 10000, 16000, 25000] ft, uniformly shifted -4.3 dB → level-flight
    # curve.  Approach = level + 3 dB (BVI uplift; ACRP02-44 lines 973-977
    # document 3-6 dB BVI rise above level flight).  Departure = level +
    # 1 dB (climb power).  Approach LOUDER than departure for helicopters —
    # opposite of fixed-wing turbofans, do not "fix" by reversing.
    "HELICOPTER": (
        [99.3, 95.9, 93.5, 91.0, 86.6, 81.2, 77.4, 72.7, 66.7, 59.9],
        [97.3, 93.9, 91.5, 89.0, 84.6, 79.2, 75.4, 70.7, 64.7, 57.9],
        100.0, 230.0, "Prop", False,
    ),
    # PC-24 placeholder (no ANP entry). Light bizjet, fuselage-mounted.
    "JET_BIZ_FUSE": (
        [98.0, 93.0, 89.0, 85.0, 78.0, 71.0, 66.0, 60.0, 54.0, 48.0],
        [101.0, 96.0, 92.0, 88.0, 81.0, 74.0, 69.0, 63.0, 57.0, 51.0],
        140.0, 320.0, "Fuselage", True,
    ),
}

# ANP installation → Rust enum name
INSTALLATION_RUST = {
    "Wing":     "Installation::Wing",
    "Fuselage": "Installation::Fuselage",
    "Prop":     "Installation::Propeller",
}

# Voronoi anchor budget per Installation (sum = NUM_CLASSES).
# Wing=9 restores B789 + A319 anchors that the EASA v9 supplement
# (commit c76cefa) had displaced when A20N/A21N gained their own NPD
# signatures. With Wing=7 the top traffic slots filled with narrow-body
# only; wide-body (B777/B787/B747/A330/A340/A350) all routed to
# WING_FALLBACK with 2.24 dB intra-class error. Bumping to 9 brings
# WING_B789 back as the wide-body anchor and recovers the original
# 0.76 dB weighted-avg per-segment error from the c93f5f4 design.
ANCHORS_PER_INSTALL: dict[str, int] = {
    "Wing":       9,
    "Fuselage":   2,
    "Prop":       2,
    "Helicopter": 1,
}
NUM_CLASSES = sum(ANCHORS_PER_INSTALL.values())  # → 14 algorithmic

# Hand-pinned anchors appended AFTER the traffic-ranked top-K selection.
# Voronoi ranks representatives by traffic volume, so rare-but-loud families
# never earn one: heavy freighters/old wide-bodies sat 3–8 dB above their
# nearest passenger anchor (B744 +8.1, B77W +4.9, MD-11 +4.4 dB mean SEL) —
# systematic under-read exactly at freight hubs, where traffic flies at night
# (Lden +10 dB weighting). Raising the Wing budget 9→10 was rejected by the
# C10b gate: traffic ranking just selects another quiet regional (E175).
# A pin only ADDS a representative — the algorithmic classes keep their
# IDENTITY (anchor + membership, minus members the pin wins), but numeric
# class indices DO re-sort by traffic afterwards; stale arrows are gated by
# the per-file contracts (airborne/cruise/airport_traffic), bumped with this
# change. Membership still flows through the same L∞ nearest-anchor pass.
# B748 (747-8, GEnx) sits mid-cluster among the heavies and pulls in
# B744/B77W/MD-11/B742/IL76 at ±0.1–2.7 dB residual; pinning B744 instead
# leaves B77W/B748/IL76 stranded (experiment 2026-06-11, audit C10b).
PINNED_ANCHORS: list[str] = ["B748"]
NUM_CLASSES += len(PINNED_ANCHORS)  # → 15

# Single source of truth for the negligible-noise designator set (sailplanes
# + hot-air balloons) — emitted into BOTH the similarity_fallback defensive
# arm and is_negligible_noise_typecode, so the two can never drift apart
# (simplifier note, 2026-06-11). BALL = ICAO 8643 special designator for
# balloon: unpowered, burner fires seconds per minute — same negligible-Lden
# argument as soaring gliders.
GLIDER_PATTERNS = [
    'b"GLID" | b"VENT" | b"DISC" | b"DUOD" | b"NIMB" | b"JANU"',
    "| [b'A', b'S', b'2', b'0'..=b'9'] | b\"AS14\" | b\"AS30\" | b\"AS31\"",
    '| b"DG40" | b"DG50" | b"DG60" | b"DG80" | b"DG1T"',
    '| b"LS8" | b"LS9" | b"LS10" | b"G103" | b"PK20" | b"BALL"',
]


@dataclass
class AnpAcft:
    acft_id: str
    npd_id: str
    engine_type: str
    num_engines: int
    weight_class: str
    lateral: str
    description: str


@dataclass
class Profile:
    typecode: str
    name: str  # display name, e.g., "B738/737800"
    approach_sel: list[float]
    departure_sel: list[float]
    approach_lmax: list[float]    # ANP LAmax NPD — peak A-weighted SPL per
    departure_lmax: list[float]   # distance row, replaces hardcoded SEL−12 dB
    v_ref_kt: float
    d_bar_m: float
    installation: str  # "Wing"/"Fuselage"/"Prop"
    proxy_source: str  # "Anp" / "Manual"
    is_jet: bool
    is_helo: bool
    # Filled during clustering:
    noise_class: int = -1   # 0..NUM_CLASSES-1, set by Voronoi assignment

    def feature(self) -> list[float]:
        """20-D NPD vector used for L∞ similarity (Voronoi).

        SEL only — Lmax is derived from same source physics and doesn't
        add independent class-discrimination signal. Keeping SEL-only
        also preserves the Tier 2.5 anchor selection.
        """
        return list(self.approach_sel) + list(self.departure_sel)


def load_aircraft(anp_dir: Path) -> dict[str, AnpAcft]:
    out: dict[str, AnpAcft] = {}
    with (anp_dir / "ANP2.3_Aircraft.csv").open() as fh:
        rdr = csv.DictReader(fh, delimiter=";")
        for row in rdr:
            acft = AnpAcft(
                acft_id=row["ACFT_ID"],
                npd_id=row["NPD_ID"],
                engine_type=row["Engine Type"],
                num_engines=int(row["Number Of Engines"]),
                weight_class=row["Weight Class"],
                lateral=row["Lateral Directivity Identifier"],
                description=row["Description"],
            )
            out[acft.acft_id] = acft
    return out


def load_v9_supplement(v9_dir: Path | None) -> tuple[dict, dict, dict]:
    """Load EASA ANP v9 supplement (April 2026) — 14 modern aircraft missing
    from v2.3 (A320neo / A321neo / A330neo / A350-1000 / 787-9 / E2 family /
    G650ER / Falcon 900EX). Returns (aircraft, sel_npd, lmax_npd) keyed by
    v9 ACFT_ID — caller merges into the v2.3 dicts (v9 wins on conflicts).

    Empty dicts when v9_dir is None — keeps the generator runnable without
    the supplement so v2.3-only regen still succeeds.
    """
    if v9_dir is None:
        return {}, {}, {}
    import openpyxl  # local import — only needed when v9 is requested
    aircraft: dict[str, AnpAcft] = {}
    sel_npd: dict[tuple[str, str], list[tuple[float, list[float]]]] = {}
    lmax_npd: dict[tuple[str, str], list[tuple[float, list[float]]]] = {}

    # Aircraft sheet
    wb = openpyxl.load_workbook(
        v9_dir / "EASA_ANP_database_Aircraft_v9.xlsx", read_only=True, data_only=True
    )
    rows = list(wb.active.iter_rows(values_only=True))
    col = {h: i for i, h in enumerate(rows[0])}
    for row in rows[1:]:
        if not row[col["ACFT_ID"]]:
            continue
        # v9 column "Wing" carries the lateral-directivity attribute
        # (Wing / Fuselage). Translate to v2.3's `Lateral Directivity
        # Identifier` semantics — same string, just different schema header.
        aircraft[row[col["ACFT_ID"]]] = AnpAcft(
            acft_id=str(row[col["ACFT_ID"]]),
            npd_id=str(row[col["NPD_ID"]]),
            engine_type=str(row[col["Engine Type"]]),
            num_engines=int(row[col["Number Of Engines"]]),
            weight_class=str(row[col["Weight Class"]]),
            lateral=str(row[col["Wing"]]),
            description=str(row[col["Description"]]),
        )
    wb.close()

    # NPD sheet
    wb = openpyxl.load_workbook(
        v9_dir / "EASA_ANP_database_NPD_Data_v9.xlsx", read_only=True, data_only=True
    )
    rows = list(wb.active.iter_rows(values_only=True))
    col = {h: i for i, h in enumerate(rows[0])}
    cols = ["L_200ft", "L_400ft", "L_630ft", "L_1000ft", "L_2000ft",
            "L_4000ft", "L_6300ft", "L_10000ft", "L_16000ft", "L_25000ft"]
    for row in rows[1:]:
        if not row[col["NPD_ID"]]:
            continue
        metric = row[col["Noise Metric"]]
        target = sel_npd if metric == "SEL" else (lmax_npd if metric == "LAmax" else None)
        if target is None:
            continue
        try:
            vals = [float(row[col[c]]) for c in cols]
            target.setdefault(
                (str(row[col["NPD_ID"]]), str(row[col["Op Mode"]])), []
            ).append((float(row[col["Power Setting"]]), vals))
        except (TypeError, ValueError):
            continue
    wb.close()
    return aircraft, sel_npd, lmax_npd


def load_npd(
    anp_dir: Path, metric: str
) -> dict[tuple[str, str], list[tuple[float, list[float]]]]:
    """(NPD_ID, Op Mode) → list of (Power Setting, [10 dB values]) for the
    given Noise Metric column. metric is one of: 'SEL', 'LAmax', 'EPNL',
    'PNLTM'. Only SEL and LAmax are consumed by the kernel."""
    out: dict[tuple[str, str], list[tuple[float, list[float]]]] = {}
    with (anp_dir / "ANP2.3_NPD_data.csv").open() as fh:
        rdr = csv.DictReader(fh, delimiter=";")
        cols = ["L_200ft", "L_400ft", "L_630ft", "L_1000ft", "L_2000ft",
                "L_4000ft", "L_6300ft", "L_10000ft", "L_16000ft", "L_25000ft"]
        for row in rdr:
            if row["Noise Metric"] != metric:
                continue
            vals = [float(row[c]) for c in cols]
            key = (row["NPD_ID"], row["Op Mode"])
            out.setdefault(key, []).append((float(row["Power Setting"]), vals))
    return out


def select_sel(npd: dict, npd_id: str, op_mode: str, power: str) -> list[float]:
    """Pick representative NPD row. power='low' for approach, 'high' for departure.
    Works for both SEL and LAmax NPD dicts (same shape)."""
    rows = npd.get((npd_id, op_mode))
    if not rows:
        return [0.0] * 10
    rows_sorted = sorted(rows, key=lambda r: r[0])
    return rows_sorted[0][1] if power == "low" else rows_sorted[-1][1]


def build_profiles(anp_dir: Path, v9_dir: Path | None = None) -> list[Profile]:
    aircraft = load_aircraft(anp_dir)
    sel_npd = load_npd(anp_dir, "SEL")
    lmax_npd = load_npd(anp_dir, "LAmax")
    # Merge EASA ANP v9 supplement (April 2026) — adds 14 modern aircraft
    # (A320neo / A321neo / A330neo / A350-1000 / 787-9 dedicated / E2 /
    # G650ER / FAL900EX). v9 ACFT_IDs win on collision; ICAO_TYPECODE_TO_ACFT_ID
    # below explicitly references v9 IDs for the typecodes that benefit.
    v9_aircraft, v9_sel, v9_lmax = load_v9_supplement(v9_dir)
    aircraft.update(v9_aircraft)
    sel_npd.update(v9_sel)
    lmax_npd.update(v9_lmax)
    profiles: list[Profile] = []

    for typecode, acft_id, manual_class in ICAO_TYPECODE_TO_ACFT_ID:
        if manual_class is not None:
            approach, departure, v_ref, d_bar, inst, is_jet = MANUAL_PROFILES[manual_class]
            # Manual profiles (HELICOPTER, GA fallbacks) have no published
            # LAmax NPD curve — use a conservative SEL−12 dB shift, matching
            # the prior hardcoded approximation. Real LAmax NPD will land
            # whenever EASA publishes rotorcraft Lmax data.
            profiles.append(Profile(
                typecode=typecode,
                name=f"{typecode}/{manual_class}",
                approach_sel=list(approach),
                departure_sel=list(departure),
                approach_lmax=[s - 12.0 for s in approach],
                departure_lmax=[s - 12.0 for s in departure],
                v_ref_kt=v_ref,
                d_bar_m=d_bar,
                installation=inst,
                proxy_source="Manual",
                is_jet=is_jet,
                is_helo=(manual_class == "HELICOPTER"),
            ))
            continue

        if acft_id not in aircraft:
            print(f"WARNING: typecode {typecode} → ACFT_ID {acft_id} not in ANP",
                  file=sys.stderr)
            continue
        acft = aircraft[acft_id]
        approach = select_sel(sel_npd, acft.npd_id, "A", "low")
        departure = select_sel(sel_npd, acft.npd_id, "D", "high")
        approach_lmax = select_sel(lmax_npd, acft.npd_id, "A", "low")
        departure_lmax = select_sel(lmax_npd, acft.npd_id, "D", "high")
        if not any(approach_lmax):
            # ANP missing LAmax for this NPD_ID — fall back to SEL−12 dB.
            approach_lmax = [s - 12.0 for s in approach]
        if not any(departure_lmax):
            departure_lmax = [s - 12.0 for s in departure]
        # v_ref / d_bar: Doc 29 standard reference values (160 kt / 370 m for
        # jets, 130 kt / 261 m for turboprops). ANP doesn't expose them
        # per-aircraft.
        if acft.engine_type == "Jet":
            v_ref, d_bar = 160.0, 370.0
        elif acft.engine_type == "Turboprop":
            v_ref, d_bar = 130.0, 261.0
        else:
            v_ref, d_bar = 90.0, 208.0
        profiles.append(Profile(
            typecode=typecode,
            name=f"{typecode}/{acft.acft_id}",
            approach_sel=approach,
            departure_sel=departure,
            approach_lmax=approach_lmax,
            departure_lmax=departure_lmax,
            v_ref_kt=v_ref,
            d_bar_m=d_bar,
            installation=acft.lateral,
            proxy_source="Anp",
            is_jet=acft.engine_type == "Jet",
            is_helo=False,
        ))

    return profiles


def recompute_fallback_npd(profiles: list[Profile], counts: dict[int, int]) -> None:
    """Replace FALLBACK NPD with traffic-weighted energy-mean of all
    non-FALLBACK profiles. Operates in-place on `profiles` for both SEL
    and LAmax NPD curves.

    Original FALLBACK NPD = B738 (proxy chosen at hand-curation time);
    that over-counts unmapped typecodes by ~2 dB at dep@1k vs the actual
    global average. Energy-mean weighted by global ADS-B segment counts
    is the unbiased estimate of "what does an unknown typecode sound
    like, given the global traffic mix observed".
    """
    fb_idx = next(i for i, p in enumerate(profiles) if p.typecode == "FALLBACK")
    others = [(i, p) for i, p in enumerate(profiles) if i != fb_idx]
    weights = [max(counts.get(i, 0), 0) for i, _ in others]
    if sum(weights) == 0:
        # No traffic data — keep the existing FALLBACK proxy values.
        return

    def energy_mean(curves: list[list[float]]) -> list[float]:
        n = len(curves[0])
        energy_sum = [0.0] * n
        weight_sum = 0.0
        for c, w in zip(curves, weights):
            if w <= 0:
                continue
            for k in range(n):
                energy_sum[k] += w * (10.0 ** (c[k] / 10.0))
            weight_sum += w
        return [10.0 * math.log10(max(e / weight_sum, 1e-30)) for e in energy_sum]

    fb = profiles[fb_idx]
    # SEL curves (approach + departure concatenated, 20-D feature)
    sel_feat = energy_mean([p.feature() for _, p in others])
    fb.approach_sel = sel_feat[:10]
    fb.departure_sel = sel_feat[10:]
    # LAmax curves — separate energy-mean since LAmax has independent
    # per-event peak distribution.
    fb.approach_lmax = energy_mean([p.approach_lmax for _, p in others])
    fb.departure_lmax = energy_mean([p.departure_lmax for _, p in others])


def linf(a: list[float], b: list[float]) -> float:
    return max(abs(x - y) for x, y in zip(a, b))


@dataclass
class AnchorClass:
    """One Voronoi class. Anchor profile is frozen — its NPD never changes."""
    anchor_idx: int                      # profile index of anchor (in profiles[])
    install: str                         # "Wing"/"Fuselage"/"Prop"/"Helicopter"
    members: list[int] = field(default_factory=list)  # all profile indices in class

    @property
    def install_rust(self) -> str:
        # "Helicopter" is a logical install label; Rust enum has only
        # Wing/Fuselage/Propeller, and helicopters use Propeller.
        return "Prop" if self.install == "Helicopter" else self.install


def cluster_voronoi(profiles: list[Profile], counts: dict[int, int]) -> list[AnchorClass]:
    """Pick top-K anchors per Installation, Voronoi-assign all members.

    Algorithm:
      1. Group profiles by sig (identical NPD vector + install + helo-flag).
      2. Sort sigs per Installation by total global traffic (descending).
      3. Anchor = top K sigs per Install (K from ANCHORS_PER_INSTALL).
      4. Each non-anchor profile → nearest anchor by L∞, restricted to
         same Installation (and helicopters → Helicopter class).

    Returns NUM_CLASSES anchor classes, ordered by total traffic descending.
    """
    # Logical install label — splits helicopters from fixed-wing Propeller.
    def install_label(p: Profile) -> str:
        return "Helicopter" if p.is_helo else p.installation

    # Group profiles by exact sig (same NPD → same sig). Members of one
    # sig collapse into a single anchor candidate; their per-typecode
    # traffic sums.
    sig_to_indices: dict[tuple, list[int]] = {}
    for i, p in enumerate(profiles):
        sig = (tuple(p.approach_sel), tuple(p.departure_sel),
               install_label(p))
        sig_to_indices.setdefault(sig, []).append(i)

    # Per Install: sort sigs by traffic, take top K as anchors.
    anchors: list[AnchorClass] = []
    used_sigs: set[tuple] = set()
    for install, k in ANCHORS_PER_INSTALL.items():
        sigs = [(sig, idxs) for sig, idxs in sig_to_indices.items()
                if sig[2] == install]
        sigs.sort(key=lambda si: -sum(counts.get(i, 0) for i in si[1]))
        for sig, idxs in sigs[:k]:
            # Anchor = highest-traffic profile within the sig (the one
            # whose typecode is most representative for naming / display).
            anchor = max(idxs, key=lambda i: counts.get(i, 0))
            anchors.append(AnchorClass(anchor_idx=anchor, install=install))
            used_sigs.add(sig)

    # Hand-pinned anchors (PINNED_ANCHORS) — appended after the ranked
    # selection so the algorithmic classes stay stable across regens.
    for tc in PINNED_ANCHORS:
        idx = next((i for i, p in enumerate(profiles) if p.typecode == tc), None)
        if idx is None:
            sys.exit(f"PINNED ANCHOR FAIL: typecode {tc!r} not in profiles")
        p = profiles[idx]
        sig = (tuple(p.approach_sel), tuple(p.departure_sel), install_label(p))
        if sig in used_sigs:
            sys.exit(f"PINNED ANCHOR FAIL: {tc} already selected by traffic "
                     f"ranking — drop it from PINNED_ANCHORS")
        anchors.append(AnchorClass(anchor_idx=idx, install=install_label(p)))
        used_sigs.add(sig)

    # Voronoi assign every profile to its nearest anchor (within same Install).
    for i, p in enumerate(profiles):
        my_install = install_label(p)
        same_install = [a for a in anchors if a.install == my_install]
        if not same_install:
            sys.exit(
                f"INTERNAL: no anchor for install={my_install!r}, "
                f"check ANCHORS_PER_INSTALL"
            )
        feat = p.feature()
        best = min(
            same_install,
            key=lambda a: linf(feat, profiles[a.anchor_idx].feature()),
        )
        best.members.append(i)

    # Sort anchors by total class traffic (descending) so emitted indices
    # roughly track popularity. Stabilises diff readability across regens.
    anchors.sort(
        key=lambda a: -sum(counts.get(m, 0) for m in a.members),
    )
    # Final assignment: write noise_class onto each Profile.
    for class_idx, a in enumerate(anchors):
        for m in a.members:
            profiles[m].noise_class = class_idx
    return anchors


def class_name(profiles: list[Profile], anchor: AnchorClass) -> str:
    """Naming: 'WING_<top_typecode>', 'FUSE_<top>', 'PROP_<top>',
    or 'HELICOPTER' for the rotorcraft class. The top typecode is the
    anchor profile's typecode; popup display will surface this name as
    the class label, so it's read by humans who recognise A320/B738/etc.
    """
    if anchor.install == "Helicopter":
        return "HELICOPTER"
    prefix = {"Wing": "WING", "Fuselage": "FUSE", "Prop": "PROP"}[anchor.install]
    return f"{prefix}_{profiles[anchor.anchor_idx].typecode}"


# Hand-tuned runway-roll reference SEL (dB at 25 m), keyed by anchor
# typecode. Anchor's own `dep@200ft` overestimates ground-roll noise by
# ~6-10 dB because flyover NPDs don't include ground absorption, engine
# baffling, or rolling-vs-overhead acoustic signature. Values calibrated
# against the pre-Tier-2 hand-tuned per-class table; deviations <3 dB
# from empirical airport measurements at 25 m. Refine when measured
# data is available. Unknown anchors fall back to `dep@200ft − 10`.
RUNWAY_DB_BY_ANCHOR: dict[str, float] = {
    "FALLBACK": 104.0,   # global energy-mean of all profiles
    "B738":     104.0,   # CFM56 narrowbody jet
    "A320":     104.0,   # IAE narrowbody jet
    "A21N":     104.0,   # neo narrowbody jet (anchor for A321 family)
    "B38M":     104.0,   # MAX narrowbody jet
    "A319":     104.0,   # narrowbody jet
    "B789":     108.0,   # widebody twin
    # Heavy widebody (pinned C10b class: B744/B748/B77W/MD11/B742/IL76).
    # Same calibration band as B789: B744/B77W previously inherited 108.0
    # via WING_B789, so their ground ops stay level while MD11/IL76 rise
    # from the 104 narrowbody band. Without this entry the dep@200ft−10
    # fallback emitted 100.1 → a −7.9 dB ground-ops regression for the
    # whole family. Raising above
    # 108 needs a cited Doc 29 / airport-measurement source.
    "B748":     108.0,
    "CRJ9":     100.0,   # regional jet, fuselage-mounted
    "C56X":      99.0,   # light bizjet, fuselage-mounted
    "C172":      92.0,   # piston SE GA
    "DH8D":      97.0,   # large turboprop
    "AS50":      94.0,   # rotorcraft (HELICOPTER class)
    # other helicopter anchors collapse to the same NPD via MANUAL_PROFILES
    "R44":       94.0,
    "EC35":      94.0,
    "EC45":      94.0,
    "B407":      94.0,
    "B06":       94.0,
}


def derive_ground_ops_per_class(profiles: list[Profile], anchors: list[AnchorClass]) -> list[list[float]]:
    """Per-class [runway, taxi, apron] per-metre LW' in dB re 1 pW/m.

    `RUNWAY_DB_BY_ANCHOR` is kept in the legacy "1 km event SEL at 25 m"
    convention so the dB values stay recognizable against ICAO Doc 29
    references. The +9.01 dB shift (= `10·log10(25/π)`) converts to the
    CNOSSOS-EU §2.5.5 per-metre formulation `recv = LW' + 10·log10(θ/d)`
    such that the identity at the 1 km / 25 m test point holds:
      recv(1km, 25m midpoint) = (anchor + 9.01) + 10·log10(θ_1km/25)
                              = anchor − 0.14 dB
    Standard Doc 29 offsets: taxi = runway − 12 dB, apron = runway − 18 dB.
    """
    lw_shift = 10.0 * math.log10(25.0 / math.pi)  # +9.01 dB
    out: list[list[float]] = []
    for a in anchors:
        anchor_p = profiles[a.anchor_idx]
        runway = RUNWAY_DB_BY_ANCHOR.get(
            anchor_p.typecode,
            anchor_p.departure_sel[0] - 10.0,
        )
        taxi = runway - 12.0
        apron = runway - 18.0
        out.append([
            round(runway + lw_shift, 2),
            round(taxi + lw_shift, 2),
            round(apron + lw_shift, 2),
        ])
    return out


def is_jet_per_class(profiles: list[Profile], anchors: list[AnchorClass]) -> list[bool]:
    return [profiles[a.anchor_idx].is_jet for a in anchors]


def emit_rust(
    profiles: list[Profile],
    anchors: list[AnchorClass],
    anp_sha: str,
    counts_sha: str,
    counts_total: int,
) -> str:
    n_profiles = len(profiles)
    n_classes = len(anchors)
    fallback_idx = n_profiles - 1  # FALLBACK is last
    fallback_class = profiles[fallback_idx].noise_class

    class_names = [class_name(profiles, a) for a in anchors]
    ground_ops = derive_ground_ops_per_class(profiles, anchors)
    is_jet = is_jet_per_class(profiles, anchors)

    lines: list[str] = []
    lines.append(f"//! Auto-generated. DO NOT EDIT BY HAND.")
    lines.append(f"//!")
    lines.append(f"//! Inputs:")
    lines.append(f"//!   ANP v2.3:        sha256={anp_sha[:16]}")
    lines.append(f"//!   global counts:   sha256={counts_sha[:16]} ({counts_total} segments)")
    lines.append(f"//!")
    lines.append(f"//! Regen: python scripts/build-aircraft-profiles.py \\")
    lines.append(f"//!         --anp <DIR> \\")
    lines.append(f"//!         --anp-v9-dir <DIR_V9> \\")
    lines.append(f"//!         --counts scripts/aircraft-profiles-counts.json \\")
    lines.append(f"//!         --ts-out frontend/src/utils/profile-class.generated.ts \\")
    lines.append(f"//!         -o engine/noise-compute/src/emission/profiles_generated.rs")
    lines.append("")
    lines.append("use super::aircraft::{Installation, NpdProfile};")
    lines.append("")
    lines.append(f"pub const NUM_PROFILES: usize = {n_profiles};")
    lines.append(f"pub const NUM_CLASSES: usize = {n_classes};")
    lines.append(f"pub const FALLBACK_PROFILE_IDX: u8 = {fallback_idx};")
    lines.append(f"pub const FALLBACK_NOISE_CLASS: u8 = {fallback_class};")
    lines.append("")
    lines.append("pub static CLASS_NAMES: [&str; NUM_CLASSES] = [")
    for n in class_names:
        lines.append(f'    "{n}",')
    lines.append("];")
    lines.append("")
    lines.append("pub static IS_JET: [bool; NUM_CLASSES] = [")
    for j, n in zip(is_jet, class_names):
        lines.append(f"    {'true' if j else 'false'}, // {n}")
    lines.append("];")
    lines.append("")
    lines.append("/// Runway-roll, taxi, apron per-metre `LW'` (dB re 1 pW/m), per noise")
    lines.append("/// class. Hand-tuned per anchor typecode (see `RUNWAY_DB_BY_ANCHOR` in")
    lines.append("/// the generator) — `dep@200ft` overestimates runway-roll by 6-10 dB")
    lines.append("/// because flyover NPDs don't include ground absorption / engine")
    lines.append("/// baffling. Standard offsets: taxi = runway − 12 dB, apron = runway")
    lines.append("/// − 18 dB. Stored values ≡ legacy 1 km event-SEL anchors + `+9.01 dB")
    lines.append("/// = 10·log10(25/π)`; identity at the 1 km test point holds under the")
    lines.append("/// CNOSSOS-EU §2.5.5 line-source receiver formula.")
    lines.append("pub static GROUND_OPS_REFERENCE_LW_PER_METER_DB: [[f64; 3]; NUM_CLASSES] = [")
    for (r, t, a), n in zip(ground_ops, class_names):
        lines.append(f"    [{r}, {t}, {a}], // {n}")
    lines.append("];")
    lines.append("")
    lines.append("/// Per-profile → noise class lookup (dense u8 index). Computed by")
    lines.append("/// Voronoi assignment to nearest anchor (L∞ on 20-D NPD vector,")
    lines.append("/// constrained to same Installation; helicopters pinned).")
    lines.append("pub static CLASS_OF_PROFILE: [u8; NUM_PROFILES] = [")
    for p in profiles:
        cls = p.noise_class
        cls_name = class_names[cls]
        lines.append(f"    {cls}, // {p.typecode} → {cls_name}")
    lines.append("];")
    lines.append("")
    lines.append("/// Anchor profile_idx for each noise class. Anchor = exact NPD vector")
    lines.append("/// of the dominant traffic profile within the class (frozen). Used by")
    lines.append("/// the kernel hot path for SEL/v_ref/d_bar/installation lookup, by the")
    lines.append("/// synth surface emitter, and by popup display name.")
    lines.append("pub static CLASS_REP_PROFILE_IDX: [u8; NUM_CLASSES] = [")
    for a, n in zip(anchors, class_names):
        anchor_p = profiles[a.anchor_idx]
        lines.append(f"    {a.anchor_idx}, // {n} → {anchor_p.typecode}")
    lines.append("];")
    lines.append("")
    lines.append("pub static PROFILES: [NpdProfile; NUM_PROFILES] = [")
    for p in profiles:
        approach = ', '.join(f'{v:.1f}' for v in p.approach_sel)
        departure = ', '.join(f'{v:.1f}' for v in p.departure_sel)
        approach_lmax = ', '.join(f'{v:.1f}' for v in p.approach_lmax)
        departure_lmax = ', '.join(f'{v:.1f}' for v in p.departure_lmax)
        lines.append(f'    NpdProfile::new(')
        lines.append(f'        "{p.name}",')
        lines.append(f'        [{approach}],')
        lines.append(f'        [{departure}],')
        lines.append(f'        [{approach_lmax}],')
        lines.append(f'        [{departure_lmax}],')
        lines.append(f'        {p.v_ref_kt},')
        lines.append(f'        {p.d_bar_m},')
        lines.append(f'        {INSTALLATION_RUST[p.installation]},')
        lines.append(f'    ),')
    lines.append("];")
    lines.append("")
    lines.append("#[inline]")
    lines.append("pub fn noise_class_of(profile_idx: u8) -> u8 {")
    lines.append("    let i = (profile_idx as usize).min(NUM_PROFILES - 1);")
    lines.append("    CLASS_OF_PROFILE[i]")
    lines.append("}")
    lines.append("")
    lines.append("/// Map ICAO 4-letter typecode to a profile index. Strict ICAO Doc 8643")
    lines.append("/// table first; on miss, falls back to `similarity_fallback` which")
    lines.append("/// pattern-matches typecode prefixes onto the closest anchor profile.")
    lines.append("/// Truly unknown / exotic typecodes still land on FALLBACK_PROFILE_IDX.")
    lines.append("pub fn profile_idx(typecode: &str) -> u8 {")
    lines.append("    match typecode {")
    for i, p in enumerate(profiles):
        if p.typecode == "FALLBACK":
            continue
        lines.append(f'        "{p.typecode}" => {i},')
    lines.append(f"        _ => similarity_fallback(typecode),")
    lines.append("    }")
    lines.append("}")
    lines.append("")
    # Similarity fallback — covers 32 % of ADSB traffic that strict-match would
    # otherwise dump into FALLBACK (per /tmp/unmapped-full.tsv 24-day scan).
    # Patterns derived from the top-1000 unmapped typecodes; all anchors named
    # here exist in the lookup table above so `profile_idx(...)` resolves.
    lines.append("/// Pattern-based fallback when the strict ICAO 4-letter match misses.")
    lines.append("/// Maps unmapped typecode prefixes onto the closest anchor profile by")
    lines.append("/// engine type / size class (Cessna piston → C172, Beech turboprop → DH8D,")
    lines.append("/// Cessna business jet → C56X, Bombardier/Gulfstream/Falcon/Hawker/Embraer")
    lines.append("/// business jet → CRJ9, Cirrus/Diamond/RV/GA-piston/UL → C172,")
    lines.append("/// helicopters → EC35). Sailplane designators route to C172 as")
    lines.append("/// defense-in-depth only — Stage 0/1 drop them up front via")
    lines.append("/// [`is_negligible_noise_typecode`].")
    lines.append("/// Typecodes that miss every pattern still return FALLBACK_PROFILE_IDX.")
    lines.append("///")
    lines.append("/// Coverage per ADSB scan (24 sample days, 1.30 M traces): patterns below")
    lines.append("/// re-route ~70–80 % of unmapped traffic — most surviving FALLBACK")
    lines.append("/// entries are exotic / military / experimental typecodes.")
    lines.append("fn similarity_fallback(typecode: &str) -> u8 {")
    lines.append("    let b = typecode.as_bytes();")
    lines.append("    // Sailplanes + self-launch/sustainer motor-gliders: Schempp-Hirth")
    lines.append("    // (Ventus/Discus/Duo Discus/Nimbus/Janus), Schleicher AS2x digit range")
    lines.append("    // (ASW-20…ASG-29; the range skips AS2T = FFA Turbine Bravo, a turboprop")
    lines.append("    // trainer) + ASK-14 / ASH-30 / ASH-31, Glaser-Dirks/DG, Rolladen-")
    lines.append("    // Schneider LS, Grob Twin G103, PIK-20E, generic GLID + BALL hot-air")
    lines.append("    // balloons. All verified against ICAO 8643 (2026-06-11). Normally dropped")
    lines.append("    // at Stage 0/1 by `is_negligible_noise_typecode`; this arm is defense-")
    lines.append("    // in-depth for any other caller and MUST precede the GL*→CRJ9,")
    lines.append("    // SF*→DH8D and helicopter arms below (GLID would otherwise evaluate as")
    lines.append("    // a CRJ-900, AS2x as an EC-135). C172 = quietest available profile,")
    lines.append("    // least-wrong if ever evaluated.")
    lines.append("    if matches!(b, " + GLIDER_PATTERNS[0])
    for pat in GLIDER_PATTERNS[1:-1]:
        lines.append("                     " + pat)
    lines.append("                     " + GLIDER_PATTERNS[-1] + ") {")
    lines.append('        return profile_idx("C172");')
    lines.append("    }")
    lines.append("    // GA piston singles (+ light piston twins — same PROP_C172 Voronoi class):")
    lines.append("    // Robin DR-400/DR-220/HR-200, Tecnam P2008/P2010, Grob G-115, Grumman")
    lines.append("    // AA-5, Mooney M20K (M20T), Socata TB-9/TB-10/TB-20, Aquila A210, MS-880")
    lines.append("    // Rallye, Commander 114, FFA AS-202 Bravo, Tecnam P2006T, Partenavia")
    lines.append("    // P-68. Each verified L1P/L2P piston in ICAO 8643 (2026-06-11).")
    lines.append("    // 2026-07-03 CZ FALLBACK scan against ICAO 8643: Zlin Z-42/142/242")
    lines.append("    // + Z-43/143 + Z-26 + Z-50 (largest real type on FALLBACK in CZ — 78")
    lines.append("    // flights/49 cells; the Novy Knin case: one Z-42 pass = 66 % of a point's")
    lines.append("    // airborne dB on the jet fallback), Cessna R182 (C82R — escapes the C1xx")
    lines.append("    // pattern), Fournier RF-6/RF-10, Jodel D150, Extra 300/NG, and Zenair")
    lines.append("    // CH-601 (CH60 — L1P fixed-wing; the CZ scan draft mislabelled it a")
    lines.append("    // helicopter).")
    lines.append("    // Previously fell to the jet-flavoured FALLBACK energy-mean — +10..25 dB")
    lines.append("    // vs reality for light pistons (audit 2026-06 airborne A1; DR40 alone =")
    lines.append("    // 1,437 rows in a 600-R4 Europe sample).")
    lines.append('    if matches!(b, b"DR40" | b"DR22" | b"HR20" | b"P208" | b"TWEN" | b"G115"')
    lines.append('                     | b"AA5" | b"M20T" | b"TB20" | b"TOBA" | b"TAMP" | b"A210"')
    lines.append('                     | b"RALL" | b"AC11" | b"AS02" | b"P06T" | b"P68"')
    lines.append('                     | b"Z42" | b"Z43" | b"Z26" | b"Z50" | b"C82R" | b"CH60"')
    lines.append('                     | b"RF6" | b"RF10" | b"D150" | b"E300" | b"EXNG") {')
    lines.append('        return profile_idx("C172");')
    lines.append("    }")
    lines.append("    // Powered ultralights / LSA (Rotax 912 class) + touring motor gliders:")
    lines.append("    // Aerospool WT9 Dynamic, Ikarus C42, generic ULAC microlight, Tecnam")
    lines.append("    // P2002 Sierra / P-92 Echo / Astore, Flight Design CT, JMB VL-3,")
    lines.append("    // Pipistrel Virus, Breezer, Evektor EuroStar + SportStar, BRM NG-5 +")
    lines.append("    // B23, PS-28 Cruiser, Sling 2, Tomark SD-4, Aero AT-3, Shark Aero,")
    lines.append("    // Direct Fly Alto, Zlin Savage, Europa; touring motor gliders HK-36")
    lines.append("    // Dimona, Scheibe SF-25 Falke, Grob G-109, Schleicher ASK-16 fly with")
    lines.append("    // engine running — UL-class noise, not sailplane silence. ICAO 8643")
    lines.append("    // special designators PARA (powered parachute) / SHIP (airship) cruise")
    lines.append("    // engine-on too — without this arm they fell to the jet FALLBACK (PARA")
    lines.append("    // only dodged it via the Piper PA** pattern below; explicit here for")
    lines.append("    // intent). GYRO is strict-mapped to HELICOPTER in the typecode table.")
    lines.append("    // 2026-07-03: ATEC Faeta (FAET) + Zephyr (ZEPH/ZEP2), TL-2000 Sting")
    lines.append("    // (TL20), Jabiru J400, Alpi Pioneer 300 (PNR3) — Czech/eur UL classics")
    lines.append("    // concentrated locally, invisible to the global-volume scan that seeded")
    lines.append("    // this list.")
    lines.append("    // Verified in ICAO 8643 (2026-06-11). Ordering: ECHO/ASTO must precede")
    lines.append("    // the EC*/AS* helicopter arm, SF25 the SF*→SAAB turboprop arm.")
    lines.append('    if matches!(b, b"WT9" | b"C42" | b"ULAC" | b"SIRA" | b"ECHO" | b"ASTO"')
    lines.append('                     | b"FDCT" | b"VL3" | b"PIVI" | b"BREZ" | b"EV97" | b"EVSS"')
    lines.append('                     | b"NG5" | b"BR23" | b"CRUZ" | b"SLG2" | b"SD4" | b"AAT3"')
    lines.append('                     | b"SHRK" | b"ALTO" | b"SAVG" | b"EUPA"')
    lines.append('                     | b"PARA" | b"SHIP"')
    lines.append('                     | b"DIMO" | b"SF25" | b"G109" | b"AS16"')
    lines.append('                     | b"FAET" | b"ZEPH" | b"ZEP2" | b"TL20" | b"J400" | b"PNR3") {')
    lines.append('        return profile_idx("C172");')
    lines.append("    }")
    lines.append("    // C-130 Hercules and C-130J Super Hercules — military 4-engine turboprop")
    lines.append("    // (would otherwise hit the C1xx Cessna piston bucket and produce a 20+ dB")
    lines.append("    //  underestimate; DH8D anchor is the closest single-class fit).")
    lines.append("    // + Shorts Skyvan (SC7, para-ops), CASA C-295 + Dornier 328 (Kbely),")
    lines.append("    // and Antonov An-2: a 1000 hp radial-piston biplane — acoustically the")
    lines.append("    // DH8D curve, not C172; sampling-window membership must not determine")
    lines.append("    // the acoustic class.")
    lines.append('    if matches!(b, b"C130" | b"C30J" | b"SC7" | b"C295" | b"D328" | b"AN2") {')
    lines.append('        return profile_idx("DH8D");')
    lines.append("    }")
    lines.append("    // Cessna piston singles (C150 / C152 / C170 / C172 / C177 / C180 / C182 / C185)")
    lines.append('    if matches!(b, [b\'C\', b\'1\', _, _]) {')
    lines.append('        return profile_idx("C172");')
    lines.append("    }")
    lines.append("    // Cessna single-engine turboprops — Caravan (C208/C20T) and Conquest (C441),")
    lines.append("    // plus CASA C-212 (also turboprop). Carved out of the C2..7xx jet range")
    lines.append("    // because the Caravan is the most common unmapped Cessna in ADSB scans.")
    lines.append('    if matches!(b, b"C208" | b"C20T" | b"C212" | b"C441") {')
    lines.append('        return profile_idx("DH8D");')
    lines.append("    }")
    lines.append("    // Cessna piston twins (C30x Skymaster, C310, C337, C340, C40x, C44x)")
    lines.append("    // and remaining piston singles (C206/C207 Stationair, C210 Centurion).")
    lines.append('    if matches!(b, [b\'C\', b\'2\', b\'0\', b\'6\' | b\'7\'] | [b\'C\', b\'2\', b\'1\', b\'0\']')
    lines.append('                     | [b\'C\', b\'3\', _, _] | [b\'C\', b\'4\', _, _]) {')
    lines.append('        return profile_idx("C172");')
    lines.append("    }")
    lines.append("    // Cessna business jets — Citation Mustang/CJ (C25x), Citation 5xx/6xx/7xx series.")
    lines.append('    if matches!(b, [b\'C\', b\'2\', b\'5\', _] | [b\'C\', b\'5\', _, _]')
    lines.append('                     | [b\'C\', b\'6\', _, _] | [b\'C\', b\'7\', _, _]) {')
    lines.append('        return profile_idx("C56X");')
    lines.append("    }")
    lines.append("    // Beechcraft Bonanza/Baron/Duke piston singles + twins (BE19/23/24/3x/5x/6x/76/77/80/88).")
    lines.append("    // Carved out of the BE turboprop bucket: BE35→DH8D was a")
    lines.append("    // 10+ dB overestimate (Bonanza is a piston single, not a King Air).")
    lines.append('    if matches!(b, [b\'B\', b\'E\', b\'1\', b\'9\']')
    lines.append('                     | [b\'B\', b\'E\', b\'2\', b\'3\' | b\'4\']')
    lines.append('                     | [b\'B\', b\'E\', b\'3\', _]')
    lines.append('                     | [b\'B\', b\'E\', b\'5\', _]')
    lines.append('                     | [b\'B\', b\'E\', b\'6\', _]')
    lines.append('                     | [b\'B\', b\'E\', b\'7\', b\'6\' | b\'7\']')
    lines.append('                     | b"BE17" | b"BE80" | b"BE88") {')
    lines.append('        return profile_idx("C172");')
    lines.append("    }")
    lines.append("    // Beechcraft King Air turboprops (BE9x / BE10 / BE20), Beechjet (BE40, small jet)")
    lines.append("    // and Beech 1900/350 commuter turboprops.")
    lines.append('    if matches!(b, [b\'B\', b\'E\', b\'9\', _]')
    lines.append('                     | b"BE10" | b"BE20" | b"BE40"')
    lines.append('                     | [b\'B\', b\'1\', b\'9\', b\'0\']')
    lines.append('                     | [b\'B\', b\'3\', b\'5\', b\'0\']) {')
    lines.append('        return profile_idx("DH8D");')
    lines.append("    }")
    lines.append("    // Boeing 717-200 (B712) — rear-fuselage twin-jet (MD-95 derivative), acoustically")
    lines.append("    // close to FUSE_CRJ9 anchor (CRJ-900 family, also rear-fuselage twin). Carved")
    lines.append("    // out before the BAe 146 / Bell helicopter B4xx pattern hits.")
    lines.append('    if matches!(b, b"B712") {')
    lines.append('        return profile_idx("CRJ9");')
    lines.append("    }")
    lines.append("    // Pilatus PC-12 / PC-7 turboprops")
    lines.append('    if matches!(b, [b\'P\', b\'C\', _, _]) {')
    lines.append('        return profile_idx("DH8D");')
    lines.append("    }")
    lines.append("    // Piper PA-2x/3x/4x — piston/turboprop singles+twins")
    lines.append('    if matches!(b, [b\'P\', b\'A\', _, _] | [b\'P\', b\'2\', b\'8\', _] | [b\'P\', b\'3\', b\'2\', _] | [b\'P\', b\'4\', b\'6\', _]) {')
    lines.append('        return profile_idx("C172");')
    lines.append("    }")
    lines.append("    // Bombardier Challenger CL-30/35/60/64 + Gulfstream/Hawker — business jet")
    lines.append("    // (`GLF*` is subsumed by [b'G', b'L', _, _] above — kept earlier as docs.)")
    lines.append('    if matches!(b, [b\'C\', b\'L\', _, _] | [b\'G\', b\'L\', _, _] | [b\'H\', b\'2\', b\'5\', _] | b"G280") {')
    lines.append('        return profile_idx("CRJ9");')
    lines.append("    }")
    lines.append("    // Falcon / Learjet / Embraer business jet (Phenom / Legacy)")
    lines.append('    if matches!(b, [b\'F\', b\'9\', b\'0\', b\'0\'] | [b\'F\', b\'2\', b\'T\', b\'H\'] | [b\'L\', b\'J\', _, _]) {')
    lines.append('        return profile_idx("CRJ9");')
    lines.append("    }")
    lines.append('    if matches!(b, [b\'E\', _, _, b\'P\' | b\'L\']) {')
    lines.append('        return profile_idx("CRJ9");')
    lines.append("    }")
    lines.append("    // Embraer EMB-110 Bandeirante / EMB-120 Brasilia — twin turboprops.")
    lines.append("    // Carved out before the E1xx jet pattern (would otherwise route to CRJ9 and")
    lines.append("    // overstate by ~12 dB SEL on small commuter turboprops).")
    lines.append('    if matches!(b, [b\'E\', b\'1\', b\'1\', _] | [b\'E\', b\'1\', b\'2\', _]) {')
    lines.append('        return profile_idx("DH8D");')
    lines.append("    }")
    lines.append("    // Embraer ERJ-1xx / E170/190/195 / E2 family (E290/E295) — regional jets.")
    lines.append('    if matches!(b, [b\'E\', b\'1\' | b\'4\' | b\'5\', _, _]) {')
    lines.append('        return profile_idx("CRJ9");')
    lines.append("    }")
    lines.append("    // ATR turboprops, Dash 8, SAAB 340/Shorts, Daher TBM")
    lines.append('    if matches!(b, [b\'A\', b\'T\', _, _] | [b\'D\', b\'H\', b\'8\', _] | [b\'S\', b\'F\', _, _] | [b\'T\', b\'B\', b\'M\', _]) {')
    lines.append('        return profile_idx("DH8D");')
    lines.append("    }")
    lines.append("    // Cirrus SR / S22, Diamond DA, Vans RV — GA piston singles")
    lines.append('    if matches!(b, [b\'S\', b\'R\', _, _] | [b\'S\', b\'2\', b\'2\', _] | [b\'D\', b\'A\', _, _] | [b\'D\', b\'V\', _, _] | [b\'R\', b\'V\', _, _]) {')
    lines.append('        return profile_idx("C172");')
    lines.append("    }")
    lines.append("    // 3-char RV* codes (RV3 / RV4 / RV6 / RV7 / RV8 / RV9 — Vans homebuilt)")
    lines.append('    if matches!(b, [b\'R\', b\'V\', _]) {')
    lines.append('        return profile_idx("C172");')
    lines.append("    }")
    lines.append("    // BAe 146 / Avro RJ — 4-engine rear-fuselage regional jet")
    lines.append("    // (carved out before the Bell B4xx helicopter pattern, which would otherwise")
    lines.append("    //  route a 4-engine jet to a single-rotor helicopter).")
    lines.append('    if matches!(b, b"B461" | b"B462" | b"B463" | b"B14R" | b"RJ70" | b"RJ85" | b"RJ1H") {')
    lines.append('        return profile_idx("CRJ9");')
    lines.append("    }")
    lines.append("    // IAI 1125 Astra / Gulfstream G100 (ASTR) — mid-size bizjet. Found during")
    lines.append("    // the ICAO 8643 audit of the former blanket AS*→helicopter arm, which")
    lines.append("    // modeled this jet as an EC-135.")
    lines.append('    if matches!(b, b"ASTR") {')
    lines.append('        return profile_idx("CRJ9");')
    lines.append("    }")
    lines.append("    // Helicopters: AS-332 Super Puma / AS-532 Cougar exact pair + EC / Bell /")
    lines.append("    // Sikorsky / Robinson / Leonardo / Airbus H-series. Deliberately NOT a")
    lines.append("    // blanket `AS**` prefix: per ICAO 8643 (verified 2026-06-11) the only")
    lines.append("    // AS-prefixed helicopter designators are AS32 / AS3B / AS50 / AS55 /")
    lines.append("    // AS65 — the latter three are strict-mapped above; every other AS* code")
    lines.append("    // is fixed-wing (Schleicher sailplanes, FFA Bravo, Tecnam Astore, IAI")
    lines.append("    // Astra — all routed above; the residue falls to FALLBACK, not to a")
    lines.append("    // phantom helicopter).")
    lines.append('    if matches!(b, b"AS32" | b"AS3B" | [b\'E\', b\'C\', _, _]) {')
    lines.append('        return profile_idx("EC35");')
    lines.append("    }")
    lines.append('    if matches!(b, [b\'B\', b\'4\', _, _] | [b\'B\', b\'5\', b\'0\', b\'5\']) {')
    lines.append('        return profile_idx("EC35");')
    lines.append("    }")
    lines.append("    // Sikorsky H-60 / S-70 / S-76 / S-92, Robinson R22/44/66, plus UHEL —")
    lines.append("    // ICAO 8643 special designator for ultralight helicopters (Mosquito")
    lines.append("    // class): piston like an R22, EC35 is the only rotorcraft anchor.")
    lines.append("    // + Mil Mi-8/17 (MI8), PZL W-3 Sokol (W3, police/army CZ), Bell 212")
    lines.append("    // (B212 — the B4xx pattern misses it), Schweizer 269 (H269), Guimbal")
    lines.append("    // Cabri G2 (G2CA, H1P piston trainer), Magni M-16 gyro (MM16 — gyros")
    lines.append("    // follow the GYRO→rotorcraft convention). Verified against ICAO 8643 2026-07-03.")
    lines.append('    if matches!(b, b"H60" | b"R22" | b"R44" | b"R66" | b"S70" | b"S76" | b"S92" | b"UHEL"')
    lines.append('                     | b"MI8" | b"W3" | b"B212" | b"H269" | b"G2CA" | b"MM16") {')
    lines.append('        return profile_idx("EC35");')
    lines.append("    }")
    lines.append("    // Leonardo AW + Airbus H-series 4-char codes")
    lines.append('    if matches!(b, [b\'A\', b\'1\', b\'0\' | b\'1\' | b\'3\' | b\'6\' | b\'8\', _]) {')
    lines.append('        return profile_idx("EC35");')
    lines.append("    }")
    lines.append('    if matches!(b, [b\'H\', b\'1\', _, _] | [b\'H\', b\'2\', b\'1\' | b\'2\', _]) {')
    lines.append('        return profile_idx("EC35");')
    lines.append("    }")
    lines.append("    FALLBACK_PROFILE_IDX")
    lines.append("}")
    lines.append("")
    lines.append("/// Beacon-only ADS-B entries that should not flow through the aircraft")
    lines.append("/// pipeline. `TWR` = control tower fixed transponders. `GND` = airport")
    lines.append("/// ground vehicles (fuel trucks, tugs, follow-me, fire trucks); these")
    lines.append("/// would otherwise hit the 737-800 fallback profile and produce a")
    lines.append("/// ~30 dB over-estimate per source at airport receivers.")
    lines.append("pub fn is_non_aircraft_typecode(typecode: &str) -> bool {")
    lines.append('    matches!(typecode.trim(), "TWR" | "GND")')
    lines.append("}")
    lines.append("")
    lines.append("/// Unpowered sailplanes + self-launch/sustainer motor-gliders whose engine")
    lines.append("/// (if any) runs minutes per flight, and hot-air balloons (BALL — burner")
    lines.append("/// blasts seconds per minute) — acoustically negligible vs every NPD")
    lines.append("/// profile we carry. Stage 0/1 drop these traces outright; the FALLBACK")
    lines.append("/// energy-mean used to model a soaring glider at 96.5 dB SEL @1000 ft —")
    lines.append("/// a jet signature (2026-06 audit, airborne A1: Halltal top-3 loudest")
    lines.append("/// included a Ventus). Code list verified against ICAO 8643 (2026-06-11):")
    lines.append("/// GLID generic glider; Schempp-Hirth VENT/DISC/DUOD/NIMB/JANU;")
    lines.append("/// Schleicher AS20–AS29 digit range (ASW/ASK/ASH/ASG families — the digit")
    lines.append("/// constraint skips AS2T, FFA's turboprop Bravo) + AS14/AS30/AS31;")
    lines.append("/// Glaser-Dirks/DG DG40/DG50/DG60/DG80/DG1T; Rolladen-Schneider")
    lines.append("/// LS8/LS9/LS10; Grob Twin G103; Eiri PIK-20E (PK20); BALL balloon.")
    lines.append("/// Touring motor gliders (DIMO/SF25/G109/AS16) are NOT here — they cruise")
    lines.append("/// engine-on and route to PROP_C172 via `similarity_fallback`.")
    lines.append("/// Blank typecodes are NOT negligible: `\"\"` stays on the FALLBACK")
    lines.append("/// energy-mean (deliberate 2026-04-29 semantics for truly-unknown types).")
    lines.append("pub fn is_negligible_noise_typecode(typecode: &str) -> bool {")
    lines.append("    matches!(")
    lines.append("        typecode.trim().as_bytes(),")
    lines.append("        " + GLIDER_PATTERNS[0])
    for pat in GLIDER_PATTERNS[1:]:
        lines.append("            " + pat)
    lines.append("    )")
    lines.append("}")
    lines.append("")

    return "\n".join(lines) + "\n"


# ─── Anchor expected values (manually pulled from ANP NPD_data.csv) ─────
# Format: typecode → (op_mode, sel_idx, expected_dB).
# Hard-coded snapshots from the ANP CSV. If the generator's CSV parsing
# regresses (wrong column extraction, off-by-one Op Mode), `verify_anchors`
# fails fast before writing the .rs file. Runtime heatmap/popup checks are
# blind to ICAO mapping fat-fingers.
# (typecode, metric "SEL"|"LAmax", op_mode, idx, expected_dB)
ANCHORS: list[tuple[str, str, str, int, float]] = [
    # B738 / 737800 / NPD CF567B / SEL D max-power (23500 lbs) col_4 (L_2000ft)
    ("B738", "SEL", "D", 4, 95.0),
    # A320 / A320-232 / NPD V2527A / SEL A min-power (2000 lbs) col_0 (L_200ft)
    ("A320", "SEL", "A", 0, 93.1),
    # B789 mapping → 7878R / NPD T1KBFP / SEL D max-power (65000 lbs) col_9 (L_25000ft)
    ("B789", "SEL", "D", 9, 68.2),
    # E190 / EMB190 / NPD CF3410E / SEL D max-power (15000 lbs) col_5 (L_4000ft)
    ("E190", "SEL", "D", 5, 84.1),
    # AT72 mapping → DHC830 / NPD PW120 / SEL D max-power (150) col_3 (L_1000ft)
    ("AT72", "SEL", "D", 3, 84.1),
    # B738 / CF567B / LAmax D max-power (23500 lbs) col_4 (L_2000ft) — guards
    # against silent ANP LAmax column parsing regressions in 1c ingestion.
    ("B738", "LAmax", "D", 4, 84.7),
]


def verify_anchors(profiles: list[Profile]) -> None:
    """Per-typecode anchor: catches CSV parsing regressions for both SEL and
    LAmax NPD column extraction. Each anchor pins ONE (metric, op_mode, idx)
    cell to its EASA ANP v2.3 published value."""
    by_tc = {p.typecode: p for p in profiles}
    for tc, metric, op_mode, idx, expected in ANCHORS:
        if tc not in by_tc:
            sys.exit(f"ANCHOR FAIL: typecode {tc!r} not in profiles list")
        p = by_tc[tc]
        if metric == "SEL":
            curve = p.departure_sel if op_mode == "D" else p.approach_sel
        else:  # LAmax
            curve = p.departure_lmax if op_mode == "D" else p.approach_lmax
        actual = curve[idx]
        if abs(actual - expected) > 0.01:
            sys.exit(
                f"ANCHOR FAIL: {tc} {metric} {op_mode}[{idx}] = {actual} dB, "
                f"expected {expected} ± 0.01.\n"
                f"  Mapped via ACFT_ID family; verify ICAO_TYPECODE_TO_ACFT_ID + "
                f"NPD_data.csv column extraction logic."
            )


def verify_voronoi(profiles: list[Profile], anchors: list[AnchorClass]) -> None:
    """Sanity: every profile assigned, every anchor in its own class,
    Voronoi distances correct.
    """
    n = len(profiles)
    if any(p.noise_class < 0 or p.noise_class >= len(anchors) for p in profiles):
        unassigned = [p.typecode for p in profiles if p.noise_class < 0]
        sys.exit(f"VORONOI FAIL: unassigned profiles: {unassigned}")
    # Each anchor must be a member of its own class.
    for ci, a in enumerate(anchors):
        if profiles[a.anchor_idx].noise_class != ci:
            sys.exit(
                f"VORONOI FAIL: anchor {profiles[a.anchor_idx].typecode} "
                f"is in class {profiles[a.anchor_idx].noise_class}, expected {ci}"
            )
    # No profile should be closer to a different same-install anchor than
    # its assigned one (within rounding tolerance).
    for p in profiles:
        feat = p.feature()
        my_anchor = anchors[p.noise_class]
        my_install = my_anchor.install
        my_dist = linf(feat, profiles[my_anchor.anchor_idx].feature())
        for ci, a in enumerate(anchors):
            if a.install != my_install or ci == p.noise_class:
                continue
            d = linf(feat, profiles[a.anchor_idx].feature())
            if d < my_dist - 1e-6:
                sys.exit(
                    f"VORONOI FAIL: {p.typecode} (class {p.noise_class}, "
                    f"L∞={my_dist:.3f}) is closer to class {ci} "
                    f"({profiles[a.anchor_idx].typecode}, L∞={d:.3f})"
                )
    # Total profiles assigned == NUM_PROFILES.
    total = sum(len(a.members) for a in anchors)
    if total != n:
        sys.exit(f"VORONOI FAIL: {total} assigned vs {n} profiles")


def emit_ts(profiles: list[Profile], anchors: list[AnchorClass]) -> str:
    """Generate `frontend/src/utils/profile-class.generated.ts`.

    Mirrors `CLASS_OF_PROFILE` + `CLASS_REP_PROFILE_IDX` from the Rust
    output so the React popup tooltip stays in sync with the Voronoi
    clustering. Without this, any anchor reshuffle (e.g. EASA v9 supplement
    promoting A20N/A21N to their own anchors) silently de-syncs the
    tooltip until a human re-types the table by hand. The CI sweep
    `python build-aircraft-profiles.py … --ts-out frontend/src/utils/profile-class.generated.ts && git diff --exit-code`
    catches drift mechanically.
    """
    class_names = [class_name(profiles, a) for a in anchors]
    lines: list[str] = []
    lines.append("// Auto-generated by scripts/build-aircraft-profiles.py — DO NOT EDIT BY HAND.")
    lines.append("//")
    lines.append("// Mirrors `CLASS_OF_PROFILE` + `CLASS_REP_PROFILE_IDX` from")
    lines.append("// `engine/noise-compute/src/emission/profiles_generated.rs`. Regenerate")
    lines.append("// alongside the Rust file via the build-aircraft-profiles.py invocation")
    lines.append("// (the `--ts-out` flag).")
    lines.append("")
    lines.append("/** ICAO typecode → noise-class label (Voronoi assignment). */")
    lines.append("export const PROFILE_CLASS: Record<string, string> = {")
    for p in profiles:
        cls_name = class_names[p.noise_class]
        lines.append(f"  {p.typecode}: '{cls_name}',")
    lines.append("}")
    lines.append("")
    lines.append("/**")
    lines.append(" * Class label → anchor profile (display typecode + whether the NPD curve")
    lines.append(" * is the synthetic traffic-weighted energy-mean of all profiles, used")
    lines.append(" * only for the FALLBACK anchor — every other class wraps a real ANP")
    lines.append(" * profile).")
    lines.append(" */")
    lines.append("export const CLASS_ANCHOR: Record<string, { typecode: string; isEnergyMean: boolean }> = {")
    for a, name in zip(anchors, class_names):
        anchor_p = profiles[a.anchor_idx]
        is_em = "true" if anchor_p.typecode == "FALLBACK" else "false"
        lines.append(f"  {name}: {{ typecode: '{anchor_p.typecode}', isEnergyMean: {is_em} }},")
    lines.append("}")
    return "\n".join(lines) + "\n"


def emit_coverage_report(profiles: list[Profile], anchors: list[AnchorClass],
                          counts: dict[int, int]) -> str:
    total = sum(counts.values()) or 1
    lines = ["", "# Coverage report", ""]
    lines.append(f"  Profiles: {len(profiles)}  Classes: {len(anchors)}  "
                 f"Global segments: {total:,}")
    lines.append("")
    by_source: dict[str, list[str]] = {"Anp": [], "Manual": []}
    for p in profiles:
        by_source.setdefault(p.proxy_source, []).append(p.typecode)
    for src, tcs in by_source.items():
        lines.append(f"  {src}: {len(tcs)} profiles")
    lines.append("")
    lines.append("# Class breakdown")
    lines.append("")
    lines.append(f"  {'#':>2} {'name':<22} {'anchor':<6} {'inst':<10} "
                 f"{'%trf':>6} {'%anchor':>7} {'#mem':>5} {'avgErr_dB':>9}")
    sigma_error = 0.0
    for ci, a in enumerate(anchors):
        anchor_p = profiles[a.anchor_idx]
        members_pct = 100 * sum(counts.get(m, 0) for m in a.members) / total
        anchor_pct = 100 * sum(
            counts.get(i, 0) for i in a.members
            if (tuple(profiles[i].approach_sel), tuple(profiles[i].departure_sel))
            == (tuple(anchor_p.approach_sel), tuple(anchor_p.departure_sel))
        ) / total
        anchor_feat = anchor_p.feature()
        # avg L∞ error weighted by traffic
        wsum, esum = 0.0, 0.0
        for m in a.members:
            w = counts.get(m, 0)
            d = linf(profiles[m].feature(), anchor_feat)
            wsum += w
            esum += w * d
        avg_err = esum / wsum if wsum > 0 else 0.0
        sigma_error += avg_err * members_pct
        cls_name = class_name(profiles, a)
        lines.append(f"  {ci:>2} {cls_name:<22} {anchor_p.typecode:<6} "
                     f"{a.install:<10} {members_pct:>5.2f}% {anchor_pct:>6.2f}% "
                     f"{len(a.members):>5} {avg_err:>8.2f}")
    lines.append("")
    lines.append(f"  Σ(avgErr × pct) = {sigma_error:.1f}    "
                 f"weighted-avg per-segment error = {sigma_error/100:.2f} dB")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--anp", required=True,
                    help="Path to extracted ANP 2.3 directory")
    ap.add_argument("--counts", required=True,
                    help="Path to global traffic counts JSON "
                         "(scripts/aircraft-profiles-counts.json)")
    ap.add_argument("--anp-v9-dir", default=None,
                    help="Optional path to extracted EASA ANP v9 supplement "
                         "(directory with EASA_ANP_database_*_v9.xlsx files)")
    ap.add_argument("-o", "--out", help="Output path (default: stdout)")
    ap.add_argument("--ts-out", default=None,
                    help="Optional path for the generated TypeScript class map "
                         "(frontend/src/utils/profile-class.generated.ts)")
    args = ap.parse_args()

    anp_dir = Path(args.anp)
    if not (anp_dir / "ANP2.3_Aircraft.csv").exists():
        sys.exit(f"ANP CSVs not found in {anp_dir}")
    v9_dir = Path(args.anp_v9_dir) if args.anp_v9_dir else None
    if v9_dir is not None and not (v9_dir / "EASA_ANP_database_Aircraft_v9.xlsx").exists():
        sys.exit(f"ANP v9 xlsx not found in {v9_dir}")
    counts_path = Path(args.counts)
    if not counts_path.exists():
        sys.exit(f"Counts JSON not found at {counts_path}")

    # Compute combined sha256 of input CSVs for provenance.
    h_anp = hashlib.sha256()
    for fname in ["ANP2.3_Aircraft.csv", "ANP2.3_NPD_data.csv"]:
        h_anp.update((anp_dir / fname).read_bytes())
    anp_sha = h_anp.hexdigest()
    counts_bytes = counts_path.read_bytes()
    counts_sha = hashlib.sha256(counts_bytes).hexdigest()

    # Counts JSON is keyed by ICAO typecode (e.g. "B738", "EMJ", "FALLBACK")
    # rather than profile_idx so a future reorder of ICAO_TYPECODE_TO_ACFT_ID
    # doesn't silently re-route counts to the wrong typecode (B738 idx 0 →
    # idx 5 swap would otherwise hand B738's 575M segments to whatever
    # typecode now sits at idx 0, and Voronoi anchors flip without any
    # failure signal). Every JSON key must resolve to a known typecode;
    # unknown keys are a fatal error so an out-of-date counts file fails
    # loud rather than silently dropping traffic.
    counts_raw = json.loads(counts_bytes)
    typecode_to_idx = {tc: i for i, (tc, _, _) in enumerate(ICAO_TYPECODE_TO_ACFT_ID)}
    unknown = [k for k in counts_raw if k not in typecode_to_idx]
    if unknown:
        sys.exit(
            f"Counts JSON has {len(unknown)} unknown typecode(s) — first 10: "
            f"{unknown[:10]}. Either add them to ICAO_TYPECODE_TO_ACFT_ID or "
            f"re-aggregate counts against the current generator."
        )
    counts: dict[int, int] = {
        typecode_to_idx[k]: int(v) for k, v in counts_raw.items()
    }
    counts_total = sum(counts.values())

    profiles = build_profiles(anp_dir, v9_dir)
    verify_anchors(profiles)

    # FALLBACK NPD must be the energy-mean BEFORE clustering, because
    # FALLBACK becomes its own anchor sig (highest single-typecode traffic
    # globally — 14.78 % — so it always wins a Wing anchor slot).
    recompute_fallback_npd(profiles, counts)

    anchors = cluster_voronoi(profiles, counts)
    if len(anchors) != NUM_CLASSES:
        sys.exit(
            f"INTERNAL: expected {NUM_CLASSES} anchors, got {len(anchors)}. "
            f"Check ANCHORS_PER_INSTALL vs available sigs per Install."
        )
    verify_voronoi(profiles, anchors)

    rust = emit_rust(profiles, anchors, anp_sha, counts_sha, counts_total)
    if args.out:
        Path(args.out).write_text(rust)
    else:
        sys.stdout.write(rust)
    if args.ts_out:
        Path(args.ts_out).write_text(emit_ts(profiles, anchors))
    sys.stderr.write(emit_coverage_report(profiles, anchors, counts))


if __name__ == "__main__":
    main()
